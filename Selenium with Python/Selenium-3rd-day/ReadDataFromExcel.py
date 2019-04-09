'''
Created on Oct 18, 2018

@author: training_d2.03.07
'''
import unittest
import os
import openpyxl



class Test(unittest.TestCase):


    def testName(self):
        path=os.getcwd()
        print(path)
        wb=openpyxl.load_workbook(r'C:\Users\training_D2.03.07\Documents\Sheet\userdetail.xlsx')
        sheetnames=wb.sheetnames
        print(sheetnames)
        sheet1=wb.active
        
        #1st way
        cellvalue=sheet1['B3'].value
        print("Cell Value from Sheet1")
        print(cellvalue)
        wb.create_sheet('Sheet4')
        # wb.save(r'C:\Users\training_D2.03.07\Documents\Sheet\userdetail.xlsx')
        # wb.remove_sheet('Sheet42')
        wb.save(r'C:\Users\training_D2.03.07\Documents\Sheet\userdetail.xlsx')
        
        celliterator=sheet1['A2':'C4']
        for c1,c2,c3 in celliterator:
            print(c1.value,c2.value,c3.value)
            
        totalrows=sheet1.max_row
        totalcolumns=sheet1.max_column
        print("Total Row and columns:", totalrows,totalcolumns)
        
        for i in range(2,totalrows+1):
            for j in range(2,totalcolumns+1):
                cell_obj=sheet1.cell(row=i,column=j)
                print(cell_obj.value)

if __name__ == "__main__":
    #import sys;sys.argv = ['', 'Test.testName']
    unittest.main()