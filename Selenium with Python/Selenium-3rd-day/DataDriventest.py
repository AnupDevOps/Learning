'''
Created on Oct 18, 2018

@author: training_d2.03.07
'''
import unittest
import openpyxl
from selenium import webdriver

class Test(unittest.TestCase):


    def testSelenium(self):
        self.driver= webdriver.Chrome(r"C:\Selenium Advance\New_chromedriver\chromedriver_win32\chromedriver.exe")
        self.driver.maximize_window()
        self.driver.get("http://localhost:8085/servlets/com.mercurytours.servlet.WelcomeServlet")
        self.driver.implicitly_wait(10)
        self.readExcel()
        # self.readExcel1()
        
        self.driver.quit()
        
        
    def Login(self,usernm,passw):
        username=self.driver.find_element_by_name("userName")
        username.send_keys(usernm)
        password=self.driver.find_element_by_name("password")
        password.send_keys(passw)
        signin=self.driver.find_element_by_name("login")
        signin.click()
        self.assertEqual(self.driver.title, "Find a Flight: Mercury Tours:", "Title page is not find")
        self.driver.find_element_by_link_text("SIGN-OFF").click()
    
    def readExcel(self):
        wb=openpyxl.load_workbook(r'C:\Users\training_D2.03.07\Documents\Sheet\userdetail.xlsx')
        sheet1=wb.active
        celliterator=sheet1['B2':'C3']
        for c1,c2 in celliterator:
            self.Login(c1.value, c2.value)
            
    def readExcel1(self):
        wb=openpyxl.load_workbook(r'C:\Users\training_D2.03.07\Documents\Sheet\userdetail.xlsx')
        sheet1=wb.active
        totalrows=sheet1.max_row
        totalcolumns=sheet1.max_column
        print("Total Row and columns:", totalrows,totalcolumns)
        
        for i in range(2,totalrows+1):
            for j in range(2,totalcolumns+1):
                cell_obj=sheet1.cell(row=i,column=j)
                print(cell_obj.value)
                self.Login(cell_obj.value[i], cell_obj.value[j])


if __name__ == "__main__":
    #import sys;sys.argv = ['', 'Test.testName']
    unittest.main()