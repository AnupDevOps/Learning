# working with Excel

import openpyxl

excelfile = openpyxl.load_workbook('example.xlsx')
print(type(excelfile))

names = excelfile.get_sheet_names()

print(names)


sheet = excelfile.get_sheet_by_name('Sheet1')
print(sheet)

sheet2 = excelfile.get_sheet_by_name('Sheet2')
print(sheet2)

print(sheet['A1'].value)
print(sheet['A2'].value)
print(sheet['A3'].value)


#print the whole data
for x in range(1,7):
    print(sheet.cell(row=x, column = 1).value)
    print(sheet.cell(row=x, column = 2).value)
    print(sheet.cell(row=x, column = 3).value)
    print("..............End of Row..........")
print("Done with reading")


# WriteExcel

import openpyxl

book = openpyxl.Workbook()
sheet=book.active

sheet['A1'] = 100
sheet['A2'] = 101
sheet['A3'] = 102

sheet['B1'] = 'Vikas'
sheet['B2'] = 'Manoo'
sheet['B3'] = 'jack'

sheet.title = "raw data"
book.save("book.xlsx")

